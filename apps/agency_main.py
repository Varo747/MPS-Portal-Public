from openpyxl import Workbook
import os
import traceback
import ctypes.wintypes

def get_desktop_path():
    CSIDL_DESKTOPDIRECTORY = 0x10
    SHGFP_TYPE_CURRENT = 0
    buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
    ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOPDIRECTORY, None, SHGFP_TYPE_CURRENT, buf)
    return buf.value

try:
    print("[STATUS] Generating dummy report", flush=True)

    desktop = get_desktop_path()
    output_path = os.path.join(desktop, "dummy_report.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws["A1"] = "This is a dummy report"
    ws["A2"] = "For test purposes only :)"
    ws["B1"] = 12345

    wb.save(output_path)

    print("[STATUS] Completed", flush=True)
    print(f"[REPORT_PATH]{output_path}", flush=True)

except Exception as e:
    print("[STATUS] ❌ Report generation failed", flush=True)
    print(traceback.format_exc(), flush=True)
