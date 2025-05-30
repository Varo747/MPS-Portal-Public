from openpyxl import Workbook
import os

print("[STATUS] Generating dummy report", flush=True)

desktop = os.path.join(os.path.expanduser("~"), "Desktop")
output_path = os.path.join(desktop, "dummy_report.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Report"

ws["A1"] = "This is a dummy report"
ws["A2"] = "For test purposes only :)"
ws["B1"] = 12345

wb.save(output_path)

print("[STATUS] Completed", flush=True)
print(f"[REPORT_PATH]{output_path}", flush=True)